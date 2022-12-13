import logging
import math


class Student:
    def __init__(self, name:str, cNum: int, code: str):
        """
        学生基本信息
        :param name: 姓名
        :param cNum: 班级编号
        :param code: 学号
        """
        self.name = name
        self.code = code
        self.cNum = cNum
        self._gender = -1
        self._history = []

    def __repr__(self):
        return str(self.cNum)+self.name + ':' + self.get_gender_cn()

    @property
    def gender(self):
        return self._gender

    @gender.setter
    def gender(self, value):
        """
        性别统一为: 0 - 女性；1 - 男性
        :param value: 目前接受 男女; man,woman; male,female
        :return:
        """
        if not isinstance(value, int):
            if 'man' == value or '男' == value or 'male' == value:
                self._gender = 1
            else:
                self._gender = 0
        else:
            if value != 0 and value != 1:
                logging.error('{} 写入性别时异常失败!'.format(self.name))
                self._gender = 0
            else:
                self._gender = value

    @property
    def history(self):
        return self._history

    @history.setter
    def history(self, value):
        if isinstance(value, list):
            self._history.extend(value)
        else:
            self._history.append(value)

    @history.getter
    def history(self):
        return self._history

    def get_gender_cn(self) -> str:
        """
        获取性别的中文名
        :return:
        """
        if self._gender == 0 or self._gender == 'female' or self._gender == 'woman' or self._gender == '女':
            return '女'
        return '男'


class Team:
    """
    小组

    """
    def __init__(self):
        self._students = [] # 本小组成员
        self.boy_num = 0
        self.girl_num = 0
        self._history = [] # 小组成员总历史
        self.member_counter = 0

    def __repr__(self):
        return '{} - {}:{}'.format(len(self._students), self.boy_num, self.girl_num)

    def __lt__(self, other):
        return self.member_counter < other.member_counter

    @property
    def students(self):
        return self._students

    @students.setter
    def students(self, value):
        if isinstance(value, list):
            self._students.extend(value)
            for student in value:
                self.add_history(student.history)
                if student.gender == 0:
                    self.girl_num += 1
                else:
                    self.boy_num += 1
                self.member_counter += 1
        else:
            self._students.append(value)
            self.add_history(value.history)
            if value.gender == 0:
                self.girl_num += 1
            else:
                self.boy_num += 1
            self.member_counter += 1

    @students.getter
    def students(self):
        return self._students

    @property
    def history(self):
        return self._history

    @history.getter
    def history(self):
        return self._history

    def add_history(self, history: list):
        if len(self._history) == 0:
            self._history = [(x,) for x in history]
        else:
            for index in range(len(history)):
                self._history[index] += (history[index],)

    def correlation_list(self, groups: list, factor, times) -> list:
        _relation = [0.00] * len(groups)
        for index in range(len(groups)):
            another = groups[index]
            relation = self.correlation(another, factor, times)
            _relation[index] = round(relation, 5)
        return _relation

    def correlation(self, another, factor: float, times: int) -> float:
        """
        计算当前小组与对比小组的关联关系
        PS: [(1,2,), (3,4,)] vs [(1,4,),(2,3,)]
        :param times: 关系最大统计伦次
        :param factor: 关系递减因子
        :param another: 进行比较的另外一个小组
        :return: 关系值
        """
        total_relation = 0.00
        denominator = len(self._history)
        if len(another.history) != denominator:
            logging.error('[{}] : [{}] 两个队伍的历史记录次数不一致，无法进行关系匹配!'.format(self, another))
            return -1
        for index in range(denominator-1, -1, -1):
            oneHander = self._history[index]
            anotherHander = another.history[index]
            repeat_time = self.getDuplicationTimes(oneHander, anotherHander)
            total_relation += (repeat_time / denominator) * math.pow(factor, denominator - 1 - index)
        return total_relation

    def getDuplicationTimes(self, one_hander:tuple, another_hander:tuple) -> int:
        result = 0
        temp_list = []
        for cur in one_hander + another_hander:
            if cur in temp_list:
                result += 1
            else:
                temp_list.append(cur)
        return result


def merge(teamList: list,
          groupTimes: int) -> list:
    """
    将队伍进行合并， 合并后队伍数量的目标就是 groupTimes的数量值。
    其中，返回出去的列表是合并后最终的队伍， 不保证能达到groupTimes数量，但是肯定不会超过；
    teamList将保留还未合并的散队

    :param teamList: 准备合并的目标队伍，最终会保留未组队的剩余成员
    :param groupTimes:
    :return:
    """
    doneList = []
    total = len(teamList)
    rest = total % groupTimes

    if total <= groupTimes:
        logging.debug('人数少于等于分组数，直接将分组本身返回即可')
        doneList = teamList.copy()
        teamList.clear()
        return doneList
    else:
        logging.debug('将无法成对的人直接作为候补人员返回， 其他人员进行组队')
        from helper.ColletionHelper import NormalTools as nt
        restList = nt.random_divide(teamList, rest)
        if int(total/groupTimes) % 2 == 1:
            doneList = nt.random_divide(teamList, groupTimes)

    while len(teamList) > groupTimes:
        curTeam = teamList.pop()
        curRelationList = curTeam.correlation_list(teamList, factor=0.7, times=6)
        _sortedRelation = sorted(curRelationList, reverse=True)
        minimal_relation = _sortedRelation.pop()
        partern_index = len(curRelationList) - curRelationList[::-1].index(minimal_relation) - 1
        partern = teamList[partern_index]

        curTeam.students = partern.students
        teamList.remove(partern)
        teamList.insert(0, curTeam)

    if int(total/groupTimes) % 2 == 1:
        for tempTeam in doneList:
            curRelationList = tempTeam.correlation_list(teamList, factor=0.7, times=6)
            _sortedRelation = sorted(curRelationList, reverse=True)
            minimal_relation = _sortedRelation.pop()
            partern_index = len(curRelationList) - curRelationList[::-1].index(minimal_relation) - 1
            partern = teamList[partern_index]

            tempTeam.students = partern.students
            teamList.remove(partern)
    else:
        doneList.extend(teamList.copy())
        teamList.clear()

    teamList.extend(restList)
    return doneList


def result_team(theTeams) -> list:
    """
    队伍拆开，成为单独一个人的初始队列
    :param theTeams: 需要呗拆解的队伍
    :return: 已经拆解好的队伍
    """
    result = []
    while theTeams:
        team = theTeams.pop()
        studentsList = team.students
        for student in studentsList:
            t = Team()
            t.students = student
            result.append(t)
    return result


if __name__ == '__main__':
    from openpyxl import load_workbook

    logging.basicConfig(level=logging.DEBUG)
    # 加载工作簿
    wb = load_workbook(r'd:\temp\HIT22VC_Grouping.xlsx')
    # 获取sheet页
    ws = wb['PreHandler']

    boys_team = []
    girls_team = []
    students = []

    """
    初始化所有同学
    记录同学的历史队伍
    同学按性别分组
    同学初始化进入队伍等待重组
    """
    for index in range(1, len(ws['C'])):
        while ws['C'][index] is None or ws['C'][index].value is None or ws['C'][index].value == '':
            break
        student = Student(name=ws['C'][index].value, code=ws['A'][index].value, cNum=ws['B'][index].value)
        student.gender = ws['D'][index].value
        students.append(student)

    # ,  'H', 'I'
    relation_columns = ['E', 'F', 'G']
    relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: [], 10:[], 11:[], 12:[], 13:[], 14:[], 15:[], 16:[], 17:[], 18:[], 19:[], 20:[]}  # 以组的维度保存组下同学所属班级编号
    for col in relation_columns:
        for index in range(1, len(ws['C'])):
            num = ws[col][index].value
            if num is None or not isinstance(num, int):
                break
            try:
                relationship.get(num).append(index)
            except AttributeError:
                print(col, index, 'ERROR')
        for groupNo, memberList in relationship.items():
            for member in memberList:
                students[member - 1].history = groupNo
        relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: [], 10:[], 11:[], 12:[], 13:[], 14:[], 15:[], 16:[], 17:[], 18:[], 19:[], 20:[]}

    # 分组前的准备工作
    # 1、将男女分开
    # 2、计算好每组男女生数量的上限、下线
    for student in students:
        t = Team()
        t.students = student
        if student.gender == 0:
            girls_team.append(t)
        else:
            boys_team.append(t)

    total = len(students)
    # boys = len(boys_team)
    # girls = len(girls_team)
    groups = 7

    # max_each = math.floor(total/groups)
    # min_each = int(total/groups)
    # max_boys = math.ceil(boys/groups)
    # min_boys = int(boys/groups)
    # max_girls = math.ceil(girls/groups)
    # min_girls = int(girls/groups)

    # 针对分好组的男女队伍进行初步分组
    # 初步分组的基本逻辑
    # 1、先将男生、女生需要额外调配的人员剥离出来
    # 2、将一定可以组合的男生组成男生基础组
    # 3、将一定可以组合的女生组成女生基础组
    # 4、将男生基础组和女生基础组进行混合
    done_boys = merge(boys_team, 8)
    done_girl = merge(girls_team, 8)

    done_boys = sorted(done_boys)
    done_girl = sorted(done_girl)
    mix_team = []
    final_team = []

    while done_boys:
        boyTeam = done_boys.pop()
        relationList = boyTeam.correlation_list(done_girl, factor=0.7, times=6)
        sortedRelation = sorted(relationList, reverse= True)
        minimal_relation = sortedRelation.pop()
        girls_index = relationList.index(minimal_relation)
        girlTeam = done_girl[girls_index]
        boyTeam.students = girlTeam.students
        mix_team.append(boyTeam)
        done_girl.remove(girlTeam)

    # 最终将额外调配的男女写入混合组， 组成最终队伍
    final_team.extend(girls_team)
    final_team.extend(boys_team)
    for final in final_team:
        if len(mix_team) <= 0:
            break
        relationList = final.correlation_list(mix_team, factor=0.7, times=6)
        sortedRelation = sorted(relationList, reverse=True)
        minimal_relation = sortedRelation.pop()
        _index = relationList.index(minimal_relation)
        _team = mix_team.pop(_index)
        final.students = _team.students

    final_team.extend(mix_team)

    # 最终打印结果的部分
    for number in range(1, total + 1):
        teamNo = 0
        founded = False
        for team in final_team:
            if founded:
                break
            teamNo += 1
            for student in team.students:
                if student.cNum == number:
                    logging.info('{}\t{}{}'.format(teamNo, student.cNum, student.name))
                    founded = True
                    break

    times = 1
    for curTeam in final_team:
        print("")
        print("#########第{}组#########".format(times))
        for student in curTeam.students:
            print('{}\t{}\t{}'.format(student.cNum, student.name, student.get_gender_cn()))
        print("")
        print("#######################")
        print("")
        times += 1

    print("DONE!")
