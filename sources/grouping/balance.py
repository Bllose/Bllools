import logging
import math
import os
import random


class Individual:
    def __repr__(self):
        return '{}-{}:{}-{}'.format(self._key, self._name, self.getGenderCN(), self.leaderCounter)

    def __init__(self):
        """
        独立的个体
        key: id
        """
        self._key = None
        self._gender = None
        self._name = None

        """
        历史记录
        """
        self._history = None
        self._leader_history = None
        self.leaderCounter = 0

    def leaderScore(self, decayList) -> float:
        '''
        计算组长得分， 每一期得分根据衰减值进行递减后想加
        :param decayList: 衰减列表
        :return: 组长统计得分
        '''
        tempList = self._leader_history.copy()
        tempList.reverse()
        decayTempList = decayList.copy()
        decayTempList.reverse()
        for index in range(len(tempList)):
            if tempList[index] == 1:
                self.leaderCounter += decayTempList[index]
        return self.leaderCounter

    @property
    def leader_history(self):
        return self._leader_history

    @leader_history.setter
    def leader_history(self, lhistory: list):
        self._leader_history = lhistory.copy()

    def set_key(self, key):
        self._key = key
        return self

    @property
    def key(self):
        return self._key

    @key.getter
    def key(self):
        return self._key

    @property
    def gender(self):
        return self._gender

    @gender.setter
    def gender(self, gender):
        self.set_gender(gender)

    @gender.getter
    def gender(self):
        return self._gender

    def set_gender(self, gender):
        if gender == '男' or gender == 'male' or gender == 'man':
            self._gender = 1
        else:
            self._gender = 0
        return self

    def getGenderCN(self):
        return '男' if self._gender == 1 else '女'

    def set_name(self, name):
        self._name = name
        return self

    @property
    def name(self):
        return self._name

    @name.getter
    def name(self):
        return self._name

    @property
    def history(self):
        return self._history

    @history.getter
    def history(self):
        return self._history

    @history.setter
    def history(self, historyList):
        """
        个人的历史记录肯定是单个数值
        但是我们后续使用的都是列表，即每个组有多个人
        所以每个人的记录也要记录为有一个元素的列表
        """
        self._history = [[x, ] for x in historyList]

    def set_history(self, history):
        self.history = history.copy()
        return self

    def set_leader_history(self, lhistory: list):
        self._leader_history = lhistory.copy()
        return self

class Group:
    def __init__(self):
        """
        队伍
        """
        self._individuals = []
        self.male = 0
        self.female = 0
        self.group_history = []
        self.leader = None # 本小组组长

    def __repr__(self):
        return '[{} : {}]'.format(self.male, self.female)

    def toString(self):
        toShowList = []
        for cur in self._individuals:
            toShowList.append(f'{cur.key}\t{cur.name}\t{cur.getGenderCN()}\t{cur.leaderCounter}')
        print('\r\n'.join(toShowList))

    def choseLeader(self, decayList: list):
        """
        根据衰减因子得到推荐组长
        :param decayList: 衰减列表
        :return:
        """
        curScore = 999999

        for person in self._individuals:
            score = person.leaderScore(decayList)
            if score < curScore:
                curScore = score
                self.leader = person

    def absorbTheLowest(self, groupList: list, decayList: list) -> int:
        """
        吸收关联关系最浅的队伍
        @param groupList: 获取最低
        @param decayList: 衰减系数列表
        """
        recorder = [0, 99999]
        for index in range(len(groupList)):
            curRelationRate = self.compareTo(groupList[index], decayList)
            if curRelationRate == 0.0:
                logging.debug(f'发现关系度为0的队伍, 直接合并！')
                recorder[0] = index
                recorder[1] = 0.0
                break
            elif curRelationRate < recorder[1]:
                logging.debug(f"发现第{index}队伍的关系度更低:[{recorder[1]}] -> [{curRelationRate}]")
                recorder[0] = index
                recorder[1] = curRelationRate
        logging.debug(f'最终确定吸收第{recorder[0]}支队伍，关系度为{recorder[1]}')
        self.individuals = groupList[recorder[0]].individuals
        groupList.remove(groupList[recorder[0]])

    def compareTo(self, target, decayList: int) -> float:
        """
        与另一个组进行比较，计算关系值
        decayList 的长度与 目标 target 下 history的长度保持一致
        @param target: 比较的另一个Group对象
        @param decay: 衰减系数
        """
        repeats = 0
        targetHistory = target.group_history
        for index in range(len(self.group_history) - 1, -1, -1):
            curHistory = self.group_history[index].copy()
            curHistory.extend(targetHistory[index]) # 将当前批次的记录混合到一起， 进行重复量统计
            temSet = set(curHistory)
            '''
            利用 set 的特性， 自动去重
            使用可能拥有重复元素的列表长度减去去重后的长度， 得到重复的总次数
            '''
            repeats += (len(curHistory) - len(temSet)) * decayList[index] # 当前轮次重复次数乘以当前轮次的衰减系数，得到当前的关系度
        return repeats

    @property
    def individuals(self):
        return self._individuals

    @individuals.setter
    def individuals(self, target):
        """
        向队伍中添加成员， 需要同步更新：
        1、性别数量
        2、总历史记录更新
        """
        if isinstance(target, list):
            if len(target) > 0:
                self._individuals.extend(target)
                for cur in target:
                    self.recordGender(cur.gender)
                    self.recordHistory(cur.history)
        else:
            self._individuals.append(target)
            self.recordGender(target.gender)
            self.recordHistory(target.history)

    @individuals.getter
    def individuals(self):
        return self._individuals

    def recordHistory(self, history: list):
        if len(self.group_history) < 1:
            # 如果是第一次记录历史记录
            # 则当前历史为空，长度为0
            for cur in history:
                self.group_history.append(cur)
        else:
            for index in range(len(self.group_history)):
                # 由于已经有历史记录了， 那么新增的历史记录应该加入原有记录中
                self.group_history[index].extend(history[index])

    def recordGender(self, gender: int):
        if gender == 1:
            self.male += 1
        else:
            self.female += 1

    @individuals.getter
    def individuals(self):
        return self._individuals


class OriginData:
    def __init__(self):
        """
        数据源对象
        本类的任务就是将数据源加载为 group对象; individual 对象。
        最终以列表的形式进行返回
        """
        self._path = None
        self._type = None
        self._file_name = None
        self._sheet = None
        self._config = None

    def set_path(self, path):
        self._path = path
        return self

    def set_type(self, type):
        self._type = type
        return self

    def set_file_name(self, name):
        self._file_name = name
        return self

    def set_sheet_name(self, sheet):
        self._sheet = sheet
        return self

    def load(self) -> list:
        """
        加载 excel, 将数据装载到 individual, group 对象中
        """
        if self._path is None or self._file_name is None or self._config is None:
            logging.error()

        from openpyxl import load_workbook
        # 加载工作簿
        wb = load_workbook(self._path + os.sep + self._file_name)
        # 获取sheet页
        ws = wb[self._sheet]

        '''
        从文件中将数据提取出来
        '''
        maxRow = ws.max_row
        # maxColumn = ws.max_column

        keyColumn = ws[self._config.key]
        nameColumn = ws[self._config.name]
        genderColumn = ws[self._config.gender]
        col_list = self._config.datas
        datasList = []
        for cur in col_list:
            curColumn = ws[cur]
            datasList.append(curColumn)

        from helper.common import color
        groups = []
        for i in range(1, maxRow):
            '''
            以行为维度，即一次处理一个人的所有信息
            首先是每次所分配组的信息
            其次是统计作为组长的信息
            '''
            from openpyxl.styles.colors import RGB
            dataList = []
            leaderList = []
            for datas in datasList:
                dataList.append(datas[i].value)
                leaderList.append(1 if datas[i].fill.fgColor.rgb == color['blue'] else 0)

            individual = Individual()
            individual\
                .set_key(keyColumn[i].value)\
                .set_name(nameColumn[i].value)\
                .set_gender(genderColumn[i].value)\
                .set_history(dataList)\
                .set_leader_history(leaderList)

            group = Group()
            group.individuals = individual
            groups.append(group)

        return groups

    def config(self):
        if self._config is None:
            self._config = self.Configure()
        return self._config

    class Configure:
        def __init__(self):
            """
            定义excel表格中的结构
            key所在列通过 key(column) 指定
            名字 name(column)
            性别 gender(column)
            数据 data(columns) 支持范围输出 -, ~ 作为分隔符

            默认第一行为表头
            如果没有表头，则使用 withOutHeader 指定
            """
            self._key = None
            self._name = None
            self._gender = None
            self._datas = []
            self._header = True

        @property
        def key(self):
            return self._key

        @key.setter
        def key(self, column):
            self._key = column

        @key.getter
        def key(self):
            return self._key

        def set_key(self, column):
            self._key = column
            return self

        @property
        def name(self):
            return self._name

        @name.setter
        def name(self, column):
            self._name = column

        @name.getter
        def name(self):
            return self._name

        def set_name(self, column):
            self._name = column
            return self

        @property
        def gender(self):
            return self._gender

        @gender.setter
        def gender(self, column):
            self._gender = column

        @gender.getter
        def gender(self):
            return self._gender

        def set_gender(self, column):
            self._gender = column
            return self

        @property
        def datas(self):
            return self._datas

        @datas.getter
        def datas(self):
            return self._datas

        @datas.setter
        def datas(self, columns):
            self.set_datas(columns)

        def set_datas(self, columns: str):
            """
            定义数据所属列
            支持范围输入，通过 - 或者 ~ 进行连接
            """
            if columns is None or len(columns) < 1:
                return self

            temp = []
            if '-' in columns:
                temp = columns.split('-')
            elif '~' in columns:
                temp = columns.split('~')

            if len(temp) == 0:
                self._datas.append(columns)
            elif len(temp) > 0:
                from helper.common import get_column_list
                self._datas.extend(get_column_list(temp[0], temp[1]))

        def withOutHeader(self):
            self._header = False
            return self


def selfMerge(target: list, order: int, decayList: list):
    """
    队列列表target自己内部进行匹配，内部合并。
    最终合并成order个组合
    @param target: 待合并的队列
    @param order: 合并最终目标
    @param decayList: 衰减列表
    """
    if target is None or len(target) < 1:
        return

    # 首先选取随机元素作为队伍选择基础
    tempList = []
    tempLen = 0
    while tempLen < order:
        chose_index = random.randint(0, len(target) - 1)
        tempList.append(target.pop(chose_index))
        tempLen = len(tempList)

    # 根据队伍基础元素，计算关联关系， 选择最低关系系数组队
    index = 0

    while len(target) > 0:
        curIndex = index % order  # 如果分7组，则0~6循环
        cur = tempList[curIndex]
        cur.absorbTheLowest(groupList=target, decayList=decayList)
        index += 1
    target.extend(tempList)


def randomDrawing(target: list, order: int) -> tuple:
    """
    随机从列表 #target 中抽取元素， 最终生成一个初步合并的队列和一个剩余队列
    @param target: 待处理列表
    @param order: 每组人数量
    """
    orderLen = len(target)
    restLen = orderLen % order
    if restLen > 0:
        # 当需要将多余列表单独分出来时
        # 通过随机数确定去除元素
        # 将随机取出人员单独保存
        restList = []
        targetLen = orderLen - restLen

        while orderLen > targetLen:
            chose_index = random.randint(0, orderLen - 1)
            restList.append(target.pop(chose_index))
            orderLen = len(target)
        return target, restList
    if orderLen >= order:
        # 如果人员数量不需要额外保存， 则直接返回对象列表， 剩余列表则为空列表
        return target, []
    else:
        # 如果总数量还不够一次完整分组， 则将所有人作为剩余部分进行返回
        return [], target


class Grouping:
    def __init__(self,
                 groupList: list,
                 decay: float = 0.7):
        """
        分组方法
        @param groupList: 待分组的原始数据，由 group 对象构成的列表
        @param decay: 衰减因子，配对时历史记录的衰减速度
        """
        self.originalDatas = groupList
        self.decayFactor = decay
        self.manList = []
        self.womanList = []
        self.finalList = []
        for group in groupList:
            if group.individuals[0].gender == 1:
                self.manList.append(group)
            else:
                self.womanList.append(group)

    def process(self, order: int):
        """
        分组的执行逻辑
        @param order: 分组数量
        """
        decayList = [math.pow(self.decayFactor, x) for x in range(len(self.manList[0].group_history))]
        decayList.reverse()

        manList, restManList = randomDrawing(self.manList, order)
        womenList, restWomenList = randomDrawing(self.womanList, order)

        # 合并 manList and womanList
        selfMerge(manList, order, decayList)
        selfMerge(womenList, order, decayList)

        # 以 manList作为基准，将 womanList 合入
        for target in manList:
            target.absorbTheLowest(groupList=womenList, decayList=decayList)

        # 将剩余的男女队伍逐一合入主列表
        restManList.extend(restWomenList)
        index = 0
        while restManList:
            cur = index % order
            manList[cur].absorbTheLowest(groupList=restManList, decayList=decayList)
            index += 1

        self.finalList = self.manList.copy()

        for team in self.finalList:
            team.choseLeader(decayList)
        return self

    def showTheLastGroup(self):
        """
        按照组的队伍输出每组人员
        :return:
        """
        print("*******************")
        print("*******************")
        index = 0
        for curGroup in self.finalList:
            curGroup.toString()
            index += 1
            if index < len(self.finalList):
                print("===================")
        print("*******************")
        print("*******************")
        return self

    def showTheLastOrder(self):
        """
        按班号输出最后排序
        :return:
        """
        orderList = [0 for i in range(len(self.originalDatas))]
        index = 1
        for curGroup in self.finalList:
            individuals = curGroup.individuals
            for individual in individuals:
                orderList[individual.key - 1] = index
            index += 1
        for position in orderList:
            print(str(position))
        return self


if __name__ == '__main__':
    od = OriginData()
    od.set_path(os.path.abspath('../../resources')) \
        .set_file_name(r'HIT22VCteam.xlsx') \
        .set_sheet_name(r'handler')
    od.config().set_key(r'B').set_name(r'C').set_gender(r'D').set_datas(r'E-V')
    grouping = Grouping(od.load())
    grouping.process(order=6).showTheLastGroup().showTheLastOrder()
