from openpyxl import load_workbook
from grouping.objects.ClassRoom import Class, Student
import logging

logging.basicConfig(level=logging.DEBUG)

# 加载工作簿
wb = load_workbook(r'..\..\resources\HIT22VCteam.xlsx')
# 获取sheet页
ws = wb['2022年上学期分组计划']

# 初始化班级信息，添加好学生信息
room = Class()
for index in range(1,len(ws['C'])):
    while ws['C'][index] is None or ws['C'][index].value is None or ws['C'][index].value == '':
        break
    student = Student(name=ws['C'][index].value, code=ws['A'][index].value, number=ws['B'][index].value)
    student.gender = ws['D'][index].value
    room.students = student

# 读取曾经的组队信息， 组装关系矩阵
relation_columns = ['E', 'F', 'G', 'H', 'I']
relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}  # 以组的维度保存组下同学所属班级编号
for col in relation_columns:
    for index in range(1,len(ws['C'])):
        num = ws[col][index].value
        if num is None or not isinstance(num, int):
            break
        try:
            relationship.get(num).append(index)
        except AttributeError:
            print(col, index, 'ERROR')
    room.add_relation(relationship)
    relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}

room.caculate_relation()
room.establish_relation_matrix()
new_relationship = room.establish_next_grouping(times=7)
print('DONE')
