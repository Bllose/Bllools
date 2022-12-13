from openpyxl import load_workbook

# 加载工作簿
wb = load_workbook(r'..\..\resources\HIT22VCteam.xlsx')
# 获取sheet页
ws = wb['2022年上学期分组计划']

total = len(ws['C'])
matrix = []
for index in range(total-1):
    matrix.append([0 for i in range(total-1)])

relation_columns = ['E', 'F', 'G', 'H', 'I']
relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}  # 以组的维度保存组下同学所属班级编号
for col in relation_columns:
    for cNum in range(1, len(ws['C'])):
        num = ws[col][cNum].value
        if num is None or not isinstance(num, int):
            break
        try:
            relationship.get(num).append(cNum)
        except AttributeError:
            print(col, cNum, 'ERROR')

    for _, group in relationship.items():
        while len(group)>0:
            key = group.pop()
            for partner in group:
                matrix[key-1][partner-1] += 1
                matrix[partner-1][key-1] += 1

    relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}


for curLine in matrix:
    curLineStr = [str(x) for x in curLine]
    print(' '.join(curLineStr))
