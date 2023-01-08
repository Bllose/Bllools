from enum import Enum
from openpyxl.styles import PatternFill


# {'red':'ffc7ce', 'yellow': 'ffeb9c', 'green': 'c6efce'}
class Color(Enum):
    RED = r'99001A'
    GREEN = r'66FF7F'
    YELLOW = r'E1FF4D'


'''
给excel cell上色
'''
final_status_fill = {'已达标': PatternFill('solid', start_color=Color.GREEN.value),
                     '未达标': PatternFill('solid', start_color=Color.RED.value),
                     '达标，但不满足0.7K': PatternFill('solid', start_color=Color.YELLOW.value),
                     '不满足0.7K': PatternFill('solid', start_color=Color.RED.value)}

'''
用来描述excel里列号
'''
all_columns = ['A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
               'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN',
               'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']


def get_column_list(start: str, end: str) -> list:
    """
    获取 excel 下的某一段列
    比如 start: 'C'; end: 'E'
    输出 ['C', 'D', 'E']
    """
    begin = -1
    stop = -1
    for i in range(len(all_columns)):
        if start == all_columns[i]:
            begin = i
        elif end == all_columns[i]:
            stop = i + 1

    if begin == -1 or stop == -1:
        return []

    return all_columns[begin: stop]
