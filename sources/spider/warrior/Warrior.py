from selenium import webdriver
import logging
from openpyxl import load_workbook
import os
from abc import ABC, abstractmethod
from openpyxl import drawing
import traceback


class Warrior(ABC):
    # root 首页：https://github.com/kcyvvf3688/djy/blob/master/gb/nf1176115.md
    def __init__(self,
                 root: str,
                 save_path: str,
                 begin: int,
                 max=2):
        logging.info('root初始化')
        self.root = root
        self.max = max
        self.counter = 0
        self.savePath = save_path
        self.begin = begin
        self.workbook = load_workbook(filename=self.findTheFile('excel-import.xlsx'))
        self.initSelenium()


    def findTheFile(self, fileName:str) -> str:
        import os
        import platform
        pathList = ['D:\\etc\\drivers\\' + fileName]
        pathList.append(os.path.join(os.getcwd(), fileName))
        if platform.system().lower() == 'windows':
            pathList.append(os.path.join(os.environ['HOMEDRIVE'] + os.environ['HOMEPATH'], fileName))
        for path in pathList:
            if os.path.exists(path):
                logging.info(path)
                return path
        logging.error('找不到文件 %s', pathList)
        input('SomeError will be occured ...')


    def initSelenium(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
        self.driver = webdriver.Chrome(executable_path=self.findTheFile('chromedriver.exe'), chrome_options=options)

    """
    总执行入口
    1、获取主页面数量
    2、遍历每一个主页面下拥有的子主题地址
    3、针对每个子主题模拟打开、截图、记录
    """
    def start(self):
        # 拿到所有主页面
        logging.info('统计页数...')
        self.getPages()

        saveFileName = os.path.join(self.savePath, "final_excel.xlsx")
        try:
            first = 1
            for curPageUrl in self.pagesUrlList:
                logging.info('从当前页面获取所有标题...')
                # 拿当前主页面下所有子主题地址
                sub_url_list = self.getSubUrlList(curPageUrl)

                for sub_url in sub_url_list:
                    if self.begin > first:
                        first += 1
                        continue
                    elif self.begin == first:
                        logging.info('已经掉过前面{}篇文章， 现在开始记录 ...'.format(str(first - 1)))
                        first += 1
                    # 针对每个子主题，截图并获取主题地址信息
                    self.recorder(sub_url)
                    if self.counter >= self.max:
                        self.workbook.save(filename=saveFileName)
                        logging.info('export: {}'.format(os.path.abspath(saveFileName)))
                        self.driver.close()
                        input('input anything to ending ...')
        except Exception as e:
            logging.error('程序异常，已经完成部分直接保存。 异常信息 {}'.format(e.args))
            traceback.print_exc()
            self.workbook.save(filename=saveFileName)
            logging.info('export: {}'.format(os.path.abspath(saveFileName)))
            self.driver.close()
            input('input anything to ending ...')



    def tryBestForUrl(self, url: str) -> bool:
        self.driver.set_page_load_timeout(15)
        try:
            self.driver.get(url)
        except:
            return True
        return False


    @abstractmethod
    def recorder(self, subUrl: str) -> None:
        """
        针对具体主题页进行解析，记录主题、地址、截图、上色等
        子类具体实现
        """
        pass


    # 获取子页面逻辑入口
    @abstractmethod
    def getSubUrlList(self, curPageUrl: str) -> list:
        """
        从当前列表页面中获取所有主题的具体地址。
        将这些主题的具体地址打包到列表中返回。
        子类具体实现。
        """
        pass


    @abstractmethod
    def getPages(self) -> list:
        """
        获取所有列表页面
        将每一页以URL的列表返回出去
        子类具体实现
        """
        pass


    def excelRecorder(self, 
                      subUrl: str, 
                      title: str,
                      host: str,
                      place: str,
                      pic: str) -> None:
        self.counter += 1
        sheet = self.workbook.active
        theRow = self.counter + 2
        sheet.cell(row=theRow, column=1).value = self.counter
        sheet.cell(row=theRow, column=2).value = title
        sheet.cell(row=theRow, column=3).value = subUrl
        sheet.cell(row=theRow, column=4).value = host
        sheet.cell(row=theRow, column=5).value = place
        sheet.row_dimensions[theRow].height = 200
        img = drawing.image.Image(pic)
        img.anchor = 'F' + str(theRow)
        sheet.add_image(img)


if __name__ == '__main__':
    warrior = Warrior(root=r'https://github.com/kcyvvf3688/djy/blob/master/gb/nf1176115.md',
                      driver_abs_path=r'D:\etc\Drivers\chromedriver.exe',
                      excel_path=r'spider\warrior\excel-import.xlsx')
    warrior.start()
