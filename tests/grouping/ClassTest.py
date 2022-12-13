from grouping.objects.ClassRoom import Class
from grouping.objects.ClassRoom import Student


class TestClass:
    def test_one(self):
        from openpyxl import load_workbook
        # 加载工作簿
        wb = load_workbook(r'..\resources\HIT22VCteam.xlsx')
        # 获取sheet页
        self.ws = wb['2022年上学期分组计划']

        # relation_columns = ['E', 'F', 'G', 'H', 'I']
        # infor_columns = ['A', 'B', 'C', 'D']

        # 初始化本班级
        # 添加学生信息
        # 做好关系的数据结构
        room = Class()
        for index in range(1,len(self.ws['C'])):
            while self.ws['C'][index] is None or self.ws['C'][index].value is None or self.ws['C'][index].value == '':
                break
            student = Student(name=self.ws['C'][index].value, code=self.ws['A'][index].value, number=self.ws['B'][index].value)
            student.gender = self.ws['D'][index].value
            room.students = student
        self.room = room
        assert room.total is len(room.students)

    def test_two(self):
        try:
            if self.room is None:
                self.test_one()
        except AttributeError:
            self.test_one()

        from grouping.processor import Divider
        relation_columns = ['E', 'F', 'G', 'H', 'I']
        relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}  # 以组的维度保存组下同学所属班级编号
        for col in relation_columns:
            for index in range(1,len(self.ws['C'])):
                num = self.ws[col][index].value
                if num is None or not isinstance(num, int):
                    break
                try:
                    relationship.get(num).append(index)
                except AttributeError:
                    print(col, index, 'ERROR')
            self.room.add_relation(relationship)
            relationship = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}

        self.room.caculate_relation()
        self.room.establish_relation_matrix()
        assert 1 is 1
