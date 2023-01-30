from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_an_excel():
    wb = Workbook()
    ws = wb.active

    data = [
        ['Apples', 10000, 5000, 8000, 6000],
        ['Pears',   2000, 3000, 4000, 5000],
        ['Bananas', 6000, 6000, 6500, 6000],
        ['Oranges',  500,  300,  200,  700],
    ]

    # add column headings. NB. these must be strings
    ws.append(["Fruit", "2011", "2012", "2013", "2014"])
    for row in data:
        ws.append(row)

    '''
    将一部分 cell 建立为表格的操作在 excel 中为 Home>Format as Table
    在 WPS 中为 开始> 表格样式> 转化为表格
    '''
    tab = Table(displayName="Table1", ref="A1:E5")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unque through out defined names and all other table name. 
    '''
    ws.add_table(tab)
    wb.save("table.xlsx")


def read_the_excel():
    from openpyxl import load_workbook
    # 加载工作簿
    wb = load_workbook("table.xlsx")
    # 获取sheet页
    ws = wb['Sheet']

    print(ws.tables)


if __name__ == '__main__':
    read_the_excel()