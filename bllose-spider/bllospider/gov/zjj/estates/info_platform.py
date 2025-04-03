import os
import logging
import re
from typing import List
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup, SoupStrainer
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bllospider.gov.zjj.estates.vos.MainTableVO import MainTableVO
from bllospider.gov.zjj.estates.constants.TableColumns import TableColumns
from bllospider.gov.zjj.estates.vos.SeatInfoVO import SeatInfoVO

# 访问路径统一前缀
PRE_URL = r'https://zjj.sz.gov.cn/ris/bol/szfdc_isz/'
# 主表格所属路径
HOST_PATH = r'index.aspx'
# 下一页按钮
NEXT_PAGE_XPATH = '/html/body/form/div[3]/div[2]/div/div[2]/div[1]/div/a[7]'
# 当前页表格体
CURRENT_PAGE_TABLE_XPATH = '//*[@id="updatepanel2"]/div/div[1]/table/tbody'
# 总条数
TOTAL_NUM = '/html/body/form/div[3]/div[2]/div/div[2]/div[2]/span'
# 进入套房信息
SUITE_INFO_ENTRY = '/html/body/form/div[4]/div/table[2]/tbody/tr[2]/td[5]/a'
# 预售栏位
SUITE_PRESALE_LAB = '/html/body/form/div[4]/div[2]/div/div[1]/ul/li[1]/a'
# 现售栏位
SUITE_SALEING_LAB = '/html/body/form/div[4]/div[2]/div/div[1]/ul/li[2]/a'
# 坐号列表
SEAT_NUMS = '//*[@id="divShowBranch"]'
# 座号信息表
SEAT_INFO_TABLE = '/html/body/form/div[4]/div[2]/div/div[3]/table/tbody'
# 套房详细信息
ROOM_DETAIL_TABLE = '//*[@id="Form1"]/div[4]/div/table/tbody'


def parse_table_row(td_tags) -> MainTableVO:
    """解析主表格行数据"""
    builder = MainTableVO.builder()
    for index, td in enumerate(td_tags):
        text = td.get_text().strip()
        href = td.find('a').get('href') if td.find('a') else ""
        
        if index == TableColumns.INDEX:
            builder.with_index(text)
        elif index == TableColumns.CERT_NUM:
            builder.with_cert(text, href, PRE_URL)
        elif index == TableColumns.PROJECT_NAME:
            builder.with_project(text, href, PRE_URL)
        elif index == TableColumns.ENTERPRICE_NAME:
            builder.with_enterprise(text)
        elif index == TableColumns.DISTRICT:
            builder.with_district(text)
        elif index == TableColumns.APPROVAL_TIME:
            builder.with_approval_time(text)
    
    return builder.build()

def fetch_main_table_data(driver, startIndex: int = 1, 
                            page_limit: int = -1, max_page: bool = False) -> List[MainTableVO]:
    """
    从列表页获取主表格数据
    Args:
        driver: Selenium WebDriver 实例
        page_limit: 爬取页面限制
        startIndex: 开始爬取的序号
    Returns:
        主表格数据列表
    """

    row_data = []
    try:
        # 打开列表页
        driver.get(PRE_URL + HOST_PATH)

        if max_page:
            # 如果想要最快速度爬取，则使用最大单页页幅，否则默认单页10条
            # 使用显式等待查找下拉框元素
            select_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(('id', 'ddlPageCount'))
            )

            # 使用 Select 类来操作下拉框
            select = Select(select_element)
            # 选择值为 20 的选项
            select.select_by_value('20')
            
            # 等待页面重新加载
            WebDriverWait(driver, 10).until(
                EC.staleness_of(select_element)
            )
        
        # 等待新的表格数据加载完成
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, '#updatepanel2 table tbody tr'))
        )
        
        total = driver.find_element(By.XPATH, TOTAL_NUM)
        match = re.search(r'\d+', total.text)
        if match:
            total_number = int(match.group())

        hasNext = True
        curPage = 1
        while hasNext:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, NEXT_PAGE_XPATH))
            )
            
            # 等待新的tbody元素出现
            tbody = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, CURRENT_PAGE_TABLE_XPATH))
            )
            
            # 获取元素的 outerHTML 属性
            tbody_html = tbody.get_attribute('outerHTML')

            if tbody_html:
                # 使用 BeautifulSoup 解析 HTML 内容
                soup = BeautifulSoup(tbody_html, 'html.parser')
                # 找到所有的 tr 标签
                tr_tags = soup.find_all('tr')

                # 遍历 tr 标签
                for tr in tr_tags:
                    logging.info("当前行的 HTML 内容：")
                    logging.info(tr.prettify())

                    th_tags = tr.find_all('th')
                    # 可以进一步提取每个 td 标签的内容
                    td_tags = tr.find_all('td')
                    if td_tags:
                        currect_vo = parse_table_row(td_tags)
                        if int(currect_vo.index) < startIndex:
                            continue
                        if currect_vo.project_path:
                            row_data.append(currect_vo)
                            if currect_vo.index == total_number:
                                hasNext = False
                                break
                        logging.info(f"当前行的数据：{row_data}")
                    elif th_tags:
                        # 表头行 跳过
                        # row_data = []
                        # for th in th_tags:
                        #     # 去除多余的空白字符
                        #     text = th.get_text().strip()
                        #     row_data.append(text)
                        # print("当前行的数据：", row_data)
                        continue    
                    else:
                        logging.warning("当前行没有 td 或 th 标签。")
                    logging.info("-" * 50)
            else:
                print("无法获取 tbody 的 outerHTML 属性。")
            # 点击下一页按钮
            element.click()
            
            # 等待新页面加载完成
            WebDriverWait(driver, 10).until(
                EC.staleness_of(element)
            )

            curPage += 1
            if page_limit > 0 and curPage > page_limit:
                hasNext = False
            # while 循环结束

    except Exception as e:
        print(f"发生错误: {e}")
    finally:
        # 关闭浏览器
        driver.quit()
    return row_data

def fetch_seat_info(driver, saleType:str, projectName:str, seatNum:str) -> List[SeatInfoVO]:
    """
    获取套房信息列表
    Args:
        driver: Selenium WebDriver 实例
        saleType: 销售类型，预售/现售
    """
    seatInfoList = []
    seatTable = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, SEAT_INFO_TABLE))
    )
    seatTableList = seatTable.get_attribute('outerHTML')
    if seatTableList:
        # 使用 BeautifulSoup 解析 HTML 内容
        soup = BeautifulSoup(seatTableList, 'html.parser')
        # 找到所有的 tr 标签
        tr_tags = soup.find_all('tr')

        curFloor = ""
        # 遍历 tr 标签
        for tr in tr_tags:
            td_tags = tr.find_all('td')
            for td in td_tags:
                roomDetailPath = td.find_all('a')
                if roomDetailPath:
                    for cur_a in roomDetailPath:
                        seatInfo = SeatInfoVO()
                        seatInfo.projectName = projectName
                        seatInfo.saleType = saleType
                        seatInfo.seatNum = seatNum
                        seatInfo.floorNum = curFloor
                        seatInfo.roomNum = td.find('div').get_text().strip()
                        seatInfo.roomStatus = cur_a.get_text().strip()
                        seatInfo.roomInfoPath = PRE_URL + cur_a.get('href')
                        seatInfoList.append(seatInfo)
                else: # 没有详情地址，说明是楼层信息
                    curFloor = td.get_text().strip()
    return seatInfoList

def fetch_seat_info_list(mainTableList:List[MainTableVO]) -> int:
    # 遍历每一个项目，梳理每个项目中的套房信息
    # 并且将套房信息插入项目对象中（MainTableVO)
    
    # 统计套房信息数量
    counter = 0
    for curRow in mainTableList:
        seatInfoList = []
        seatMap = {}
        projectName = curRow.project_name
        projectPath = curRow.project_path

        # 项目套房信息页面
        driver = webdriver.Chrome(service=service)
        driver.get(projectPath)
        suiteInfoEntryElement = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, SUITE_INFO_ENTRY))
        )
        suiteInfoEntryElement.click()
        # 等待新页面加载完成    
        WebDriverWait(driver, 10).until(
            EC.staleness_of(suiteInfoEntryElement)
        )

        # 等待座号加载完成
        seatNums = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, SEAT_NUMS))
        )
        # 获取元素的 outerHTML 属性
        seatNumList = seatNums.get_attribute('outerHTML')

        if seatNumList:
            # 使用 BeautifulSoup 解析 HTML 内容
            soup = BeautifulSoup(seatNumList, 'html.parser')
            a_tags = soup.find_all('a')
            for cur_a in a_tags:
                seatName = cur_a.get_text().strip()
                seatPath = cur_a.get('href')
                seatMap[seatName] = PRE_URL + seatPath
        driver.close()

        for seatNum, value in seatMap.items():
            print(f"Key: {seatNum}, Value: {value}")
            seatInfoDriver = webdriver.Chrome(service=service)
            seatInfoDriver.get(value)

            # 预售页/现售页
            # 等待新页面加载完成并确保进入预售页面
            suitePresaleElement = WebDriverWait(seatInfoDriver, 10).until(
                EC.presence_of_element_located((By.XPATH, SUITE_PRESALE_LAB))
            )
            suitePresaleElement.click()
            # 等待新页面加载完成    
            WebDriverWait(seatInfoDriver, 10).until(
                EC.staleness_of(suitePresaleElement)
            )
            # 再当前页面重新找到当前处理座号
            target_a_tag = WebDriverWait(seatInfoDriver, 10).until(
                EC.presence_of_element_located((By.XPATH, f'//a[text()="{seatNum}"]'))
            )
            # 再次点击座号进入页面
            target_a_tag.click()

            seatInfoList.extend(fetch_seat_info(seatInfoDriver, '预售', projectName, seatNum))

            suiteSalingElement = WebDriverWait(seatInfoDriver, 10).until(
                EC.presence_of_element_located((By.XPATH, SUITE_SALEING_LAB))
            )
            suiteSalingElement.click()
            # 等待新页面加载完成    
            WebDriverWait(seatInfoDriver, 10).until(
                EC.staleness_of(suiteSalingElement)
            )
            # 再当前页面重新找到当前处理座号
            target_a_tag = WebDriverWait(seatInfoDriver, 10).until(
                EC.presence_of_element_located((By.XPATH, f'//a[text()="{seatNum}"]'))
            )
            # 再次点击座号进入页面
            target_a_tag.click()
            seatInfoList.extend(fetch_seat_info(seatInfoDriver, '现售', projectName, seatNum))
            seatInfoDriver.close()
        
        curRow.seat_info_list = seatInfoList
        counter += len(seatInfoList)
    return counter

def fetch_room_detail_info(mainTableList:List[MainTableVO]) -> None:
    roomInfoDriver = webdriver.Chrome(service=service)
    for curRow in mainTableList:
        logging.info(curRow)
        for curSeatInfo in curRow.seat_info_list:
            logging.info(curSeatInfo)
            roomInfoPath = curSeatInfo.roomInfoPath
            
            roomInfoDriver.get(roomInfoPath)
            roomInfoTableElement = WebDriverWait(roomInfoDriver, 10).until(
                EC.presence_of_element_located((By.XPATH, ROOM_DETAIL_TABLE))
            )
            room_detail_table_html = roomInfoTableElement.get_attribute('outerHTML')
            if room_detail_table_html:
                # 使用 BeautifulSoup 解析 HTML 内容
                soup = BeautifulSoup(room_detail_table_html, 'html.parser')
                # 找到所有的 tr 标签
                tr_tags = soup.find_all('tr')

                rowOne = tr_tags[0]
                rowOneList = rowOne.find_all('td')
                curSeatInfo.projectBuildingInfo = rowOneList[1].get_text().strip()

                rowTwo = tr_tags[1]
                rowTwoList = rowTwo.find_all('td')
                curSeatInfo.use = rowTwoList[3].get_text().strip()
                curSeatInfo.isAccessibility = rowTwoList[5].get_text().strip()

                rowThree = tr_tags[2]
                rowThreeList = rowThree.find_all('td')
                curSeatInfo.price = rowThreeList[1].get_text().strip()
                curSeatInfo.price2 = rowThreeList[2].get_text().strip()
                curSeatInfo.contractNum = rowThreeList[4].get_text().strip()

                rowFive = tr_tags[4]
                rowFiveList = rowFive.find_all('td')
                curSeatInfo.area = rowFiveList[1].get_text().strip()
                curSeatInfo.area2 = rowFiveList[3].get_text().strip()
                curSeatInfo.area3 = rowFiveList[5].get_text().strip()

                rowSeven = tr_tags[6]
                rowSevenList = rowSeven.find_all('td')
                curSeatInfo.area4 = rowSevenList[1].get_text().strip()
                curSeatInfo.area5 = rowSevenList[3].get_text().strip()
                curSeatInfo.area6 = rowSevenList[5].get_text().strip()
    roomInfoDriver.close()            


if __name__ == '__main__':
    # 获取当前脚本所在的目录
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建 chromedriver 的路径
    chrome_driver_path = os.path.join(script_dir, 'chromedriver.exe')

    # 创建 ChromeDriver 服务对象
    service = Service(chrome_driver_path)

    # 创建 Chrome 浏览器实例
    driver = webdriver.Chrome(service=service)

    # 获取主列表数据，同时控制数据获取范围
    mainTableList = fetch_main_table_data(driver, startIndex = 1, page_limit= 1)
    logging.info(f"主表信息获取完毕，信息总条数：{len(mainTableList)}")

    counter = fetch_seat_info_list(mainTableList)
    logging.info(f"套房信息获取完毕，信息总条数: {counter}")

    try:
        fetch_room_detail_info(mainTableList)
        logging.info(f"房间详情信息获取完毕")
    except Exception as e:
        logging.error(f"获取房间详情信息发生错误: {e}")
        # 无论报了什么错，将已经处理好的信息进行保存

    # 创建一个新的工作簿
    workbook = Workbook()

    # 删除默认创建的工作表
    if 'Sheet' in workbook.sheetnames:
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)
    
    for curRow in mainTableList:
        # 新建一个工作表并命名
        new_sheet_name = curRow.project_name
        new_sheet = workbook.create_sheet(title=new_sheet_name)

        # 写入表头
        new_sheet.append(["预售证号", "项目名称", "开发企业", "所在区", "批准时间", 
        "预售/现售", "座号", "层号", "房间号", "出售状态", "项目楼栋情况", "用途", 
        "是否无障碍住房", "拟售价格（按建筑面积计）", "拟售价格（按套内建筑面积计）", 
        "合同编号", "预售查丈建筑面积", "预售查丈套内建筑面积", "预售查丈分摊面积",
        "竣工查丈建筑面积", "竣工查丈套内建筑面积", "竣工查丈分摊面积"])

        for curSeatInfo in curRow.seat_info_list:
            new_sheet.append([curRow.cert_num, curRow.project_name, curRow.enterprice_name, curRow.district, curRow.approval_time, 
            curSeatInfo.saleType, curSeatInfo.seatNum, curSeatInfo.floorNum, curSeatInfo.roomNum, curSeatInfo.roomStatus, curSeatInfo.projectBuildingInfo, curSeatInfo.use,
            curSeatInfo.isAccessibility, curSeatInfo.price, curSeatInfo.price2, 
            curSeatInfo.contractNum, curSeatInfo.area, curSeatInfo.area2, curSeatInfo.area3,
            curSeatInfo.area4, curSeatInfo.area5, curSeatInfo.area6])
    
    # 保存工作簿为 xlsx 文件
    file_path = "output.xlsx"
    workbook.save(file_path)
    print(f"数据已成功写入 {file_path} 中的 {new_sheet_name} 工作表")





