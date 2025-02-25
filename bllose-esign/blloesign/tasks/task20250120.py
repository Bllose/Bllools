from blloesign.esign.Client import eqb_sign
from openpyxl import load_workbook



def handler(workbook, sheetName:str, env:str = 'pro'):
    client = eqb_sign(env=env)
    print(f'----------> SHEET: {sheetName}')
    sheet = workbook[sheetName]
    for row in sheet.iter_rows(min_row=2, values_only=False):
        # 获取 A 列和 B 列的值
        orderNo = row[2].value
        flowId = row[1].value
        response_json = client.getSignFlowDetail(signFlowId=flowId)

        signers = response_json['data']['signers']
        for signer in signers:
            signerType = signer['signerType']
            if signerType == 1:
                signFields = signer['signFields']
                for signField in signFields:
                    normalSignFieldConfig = signField['normalSignFieldConfig']
                    sealId = normalSignFieldConfig['sealId']
                    sealOwnerId = normalSignFieldConfig['sealOwnerId']
                    row[3].value = sealOwnerId
                    row[4].value = sealId
                    print(f'{orderNo}\t{flowId}\t{sealId}\t{sealOwnerId}')

def analysis_the_data(abs_path:str, sheetNameList:list, env:str = 'pro'):
    workbook = load_workbook(abs_path, env=env)
    [handler(workbook, sheetName) for sheetName in sheetNameList]
    workbook.save('output.xlsx')


if __name__ == '__main__':
    # workbook = load_workbook(r'C:\Users\bllos\Desktop\待检查数据.xlsx')

    # handler(workbook, '检查是否盖了惠州TCL')
    # handler(workbook, '泰安沛辉太阳能发电有限公司')
    # workbook.save('output.xlsx')

    analysis_the_data(r'C:\Users\bllos\Desktop\待检查数据.xlsx', ['检查是否盖了惠州TCL', '泰安沛辉太阳能发电有限公司'])

    


    